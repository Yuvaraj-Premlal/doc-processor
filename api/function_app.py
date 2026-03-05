"""
function_app.py — Azure Functions backend for PDF bundle processing.

Pipeline per uploaded PDF:
  upload  →  [queue]  →  job_worker:
      1. classify   (DI classifier — full PDF, sequential)
      2. extract    (DI models — one per doc type, PARALLEL)
      3. save       (Cosmos DB)

Stage tracking written to Cosmos on every transition so clients can poll
GET /job/{jobId}/status in real time.

Key design decisions:
  - Queue is NEVER blindly cleared on upload.  Instead every message carries a
    batchId.  Workers discard messages whose batchId doesn't match the latest
    active batch recorded in Cosmos.  This eliminates the race condition where
    clear_messages() wiped freshly-enqueued jobs.
  - A single PdfReader is created per job and shared (read-only) across threads.
  - Each parallel extraction thread creates its own DI client (SDK not thread-safe).
  - Progress helpers (_set_stage, _set_extraction_substatus) NEVER raise —
    they log warnings and return so the main pipeline is never disrupted.
  - Every public helper is typed and documented.
"""

import logging
import azure.functions as func
import json
import os
import uuid
import base64
import io
import zipfile
import concurrent.futures
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Tuple

from azure.storage.blob import BlobServiceClient
from azure.storage.queue import QueueClient
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.cosmos import CosmosClient
from azure.cosmos.exceptions import CosmosResourceNotFoundError
from pypdf import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)


# ════════════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════════════
UPLOADS_CONTAINER      = os.environ.get("UPLOADS_CONTAINER", "uploads")
COSMOS_DATABASE        = os.environ.get("COSMOS_DATABASE", "pdfbundle")
COSMOS_CONTAINER       = os.environ.get("COSMOS_CONTAINER", "results")
CLASSIFIER_ID          = os.environ.get("DI_CLASSIFIER_ID", "cevadocclassmodel")
ACTIVE_BATCH_DOC_ID    = "__active_batch__"   # Cosmos sentinel doc — never returned to callers

TARGET_DOC_TYPES = ["CEVA", "ENTRY SUMMARY", "PARTS WORKSHEET"]

EXTRACTION_MODELS: Dict[str, str] = {
    "CEVA":            "ceva_invoice_model",
    "ENTRY SUMMARY":   "entry-summary-v1",
    "PARTS WORKSHEET": "partsworksheet_model",
}

DEFAULT_PAGE_SIZE       = 20
MAX_PAGE_SIZE           = 200
CEVA_INVOICE_DATE_FIELD = "INVOICE DATE"

# ── DI timeout / retry ────────────────────────────────────────
# Azure Functions queue trigger visibility timeout is 10 min (600 s).
# Keep total DI work well under that so messages are never re-delivered
# while still processing. Adjust via app settings if your PDFs are large.
DI_CLASSIFY_TIMEOUT_SEC = int(os.environ.get("DI_CLASSIFY_TIMEOUT_SEC", "120"))
DI_EXTRACT_TIMEOUT_SEC  = int(os.environ.get("DI_EXTRACT_TIMEOUT_SEC",  "120"))
DI_MAX_RETRIES          = int(os.environ.get("DI_MAX_RETRIES",          "2"))

# ── Stage constants ───────────────────────────────────────────
STAGE_QUEUED      = "queued"
STAGE_CLASSIFYING = "classifying"
STAGE_EXTRACTING  = "extracting"
STAGE_SAVING      = "saving"
STAGE_COMPLETED   = "completed"
STAGE_FAILED      = "failed"
STAGE_DISCARDED   = "discarded"

# Per-doc-type extraction sub-status
SUB_PENDING    = "pending"
SUB_PROCESSING = "processing"
SUB_DONE       = "done"
SUB_FAILED     = "failed"


# ════════════════════════════════════════════════════════════════
# TIME HELPERS
# ════════════════════════════════════════════════════════════════
def utc_now() -> datetime:
    return datetime.now(timezone.utc)

def utc_now_iso() -> str:
    return utc_now().isoformat()


# ════════════════════════════════════════════════════════════════
# CLIENT FACTORIES
# Each call returns a fresh client — safe for Azure Functions
# cold/warm-start behaviour and avoids stale connection pools.
# ════════════════════════════════════════════════════════════════
def _blob_service() -> BlobServiceClient:
    return BlobServiceClient.from_connection_string(os.environ["AzureWebJobsStorage"])

def _queue_client() -> QueueClient:
    """Return a QueueClient, ensuring the queue exists."""
    q = QueueClient.from_connection_string(os.environ["AzureWebJobsStorage"], "jobs")
    try:
        q.create_queue()
    except Exception:
        pass  # Already exists — fine
    return q

def _di_client() -> DocumentIntelligenceClient:
    return DocumentIntelligenceClient(
        endpoint=os.environ["DI_ENDPOINT"],
        credential=AzureKeyCredential(os.environ["DI_KEY"]),
    )

def _cosmos_container():
    client = CosmosClient(
        os.environ["COSMOS_ENDPOINT"],
        credential=os.environ["COSMOS_KEY"],
    )
    return client.get_database_client(COSMOS_DATABASE).get_container_client(COSMOS_CONTAINER)


# ════════════════════════════════════════════════════════════════
# DI CALL WRAPPER — timeout + retry
#
# DI pollers (.result()) block indefinitely by default.
# We run them in a ThreadPoolExecutor so we can impose a hard timeout.
# On timeout or transient error we retry up to DI_MAX_RETRIES times
# with a short back-off, then raise so the job is marked failed
# (not silently stuck forever at "classifying").
# ════════════════════════════════════════════════════════════════
import time as _time

def _di_call_with_retry(
    fn,           # callable that starts the DI operation and returns a poller
    timeout_sec: int,
    label: str,   # e.g. "classify" or "extract CEVA" — used in log messages
) -> Any:
    """
    Call fn() to get a DI poller, then wait up to timeout_sec for the result.
    Retries up to DI_MAX_RETRIES times on any exception or timeout.
    Raises RuntimeError if all attempts fail — caller records STAGE_FAILED.
    """
    last_exc: Optional[Exception] = None

    for attempt in range(1, DI_MAX_RETRIES + 2):   # +2: first attempt + retries
        try:
            logger.info("[DI] %s  attempt=%d/%d", label, attempt, DI_MAX_RETRIES + 1)
            poller = fn()

            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                future = ex.submit(poller.result)
                try:
                    return future.result(timeout=timeout_sec)
                except concurrent.futures.TimeoutError:
                    raise TimeoutError(
                        f"DI call '{label}' timed out after {timeout_sec}s "
                        f"(attempt {attempt})"
                    )

        except Exception as exc:
            last_exc = exc
            logger.warning("[DI] %s  attempt=%d FAILED: %s", label, attempt, exc)
            if attempt <= DI_MAX_RETRIES:
                backoff = 5 * attempt
                logger.info("[DI] Retrying %s in %ds...", label, backoff)
                _time.sleep(backoff)

    raise RuntimeError(
        f"DI call '{label}' failed after {DI_MAX_RETRIES + 1} attempts. "
        f"Last error: {last_exc}"
    )


# ════════════════════════════════════════════════════════════════
# SMALL UTILS
# ════════════════════════════════════════════════════════════════
def _normalize_doctype(s: Optional[str]) -> str:
    return (s or "").strip().upper()

def _safe_slug(s: str) -> str:
    return (s or "").strip().upper().replace(" ", "_")

def _is_pdf(name: str) -> bool:
    return (name or "").lower().endswith(".pdf")

def _is_zip(name: str) -> bool:
    return (name or "").lower().endswith(".zip")

def _parse_int(s: Optional[str], default: int) -> int:
    try:
        return int(s)
    except Exception:
        return default

def _parse_iso_date_only(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    s = s.strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None

def _delete_blob_prefix(container_client, prefix: str) -> None:
    for blob in container_client.list_blobs(name_starts_with=prefix):
        try:
            container_client.delete_blob(blob.name)
        except Exception as exc:
            logger.warning("Could not delete blob %s: %s", blob.name, exc)

def _json_response(data: Any, status_code: int = 200) -> func.HttpResponse:
    return func.HttpResponse(
        json.dumps(data, default=str),
        status_code=status_code,
        mimetype="application/json",
    )


# ════════════════════════════════════════════════════════════════
# BATCH MANAGEMENT
#
# Design: we never blindly clear the queue.
# Every queued message carries a batchId.
# On upload we write the new batchId to a Cosmos sentinel doc.
# Workers read that sentinel and silently discard any message
# whose batchId does not match — handling the race where a message
# was already invisible (dequeued) when clear_messages() would have run.
# ════════════════════════════════════════════════════════════════
def _set_active_batch(batch_id: str) -> None:
    """
    Persist the current active batch ID to Cosmos.
    Raises on failure — upload should be aborted if this cannot be written,
    since without it the stale-message guard cannot function.
    """
    _cosmos_container().upsert_item({
        "id":      ACTIVE_BATCH_DOC_ID,
        "type":    "active_batch",
        "batchId": batch_id,
        "setAt":   utc_now_iso(),
    })
    logger.info("[BATCH] Active batch set → %s", batch_id)


def _get_active_batch_id() -> Optional[str]:
    """Read the current active batch ID from Cosmos. Returns None on any failure."""
    try:
        doc = _cosmos_container().read_item(
            item=ACTIVE_BATCH_DOC_ID,
            partition_key=ACTIVE_BATCH_DOC_ID,
        )
        return doc.get("batchId")
    except CosmosResourceNotFoundError:
        return None
    except Exception as exc:
        logger.warning("[BATCH] Could not read active batch: %s", exc)
        return None


def _clear_stale_queue_messages() -> None:
    """
    Delete pending (visible) messages from the queue — best effort.
    Only clears when messages actually exist to avoid spurious log noise.
    Messages already dequeued by a worker are handled by the batchId check.
    Never raises — failure here is non-fatal.
    """
    try:
        qc    = _queue_client()
        props = qc.get_queue_properties()
        count = props.get("approximate_message_count", 0)
        if count and count > 0:
            qc.clear_messages()
            logger.info("[QUEUE] Cleared %d stale message(s)", count)
        else:
            logger.info("[QUEUE] Queue already empty — no clear needed")
    except Exception as exc:
        logger.warning("[QUEUE] Could not clear queue (non-fatal): %s", exc)


# ════════════════════════════════════════════════════════════════
# PROGRESS TRACKING
#
# _set_stage and _set_extraction_substatus NEVER raise.
# They log a warning and return so the main pipeline is unaffected.
#
# Cosmos document shape while processing:
# {
#   "id":        "<jobId>",
#   "type":      "bundle_result",
#   "fileName":  "shipment.pdf",
#   "batchId":   "<batchId>",
#   "status":    "processing",       <- coarse (for backwards compat)
#   "stage":     "extracting",       <- fine-grained current stage
#   "stageAt":   "<iso>",
#   "stagelog":  [
#     {"stage": "queued",      "at": "..."},
#     {"stage": "classifying", "at": "..."},
#     {"stage": "extracting",  "at": "..."}
#   ],
#   "extraction": {
#     "CEVA":            {"status": "done",       "startedAt": "...", "completedAt": "...", "fieldCount": 12},
#     "ENTRY SUMMARY":   {"status": "processing", "startedAt": "...", "modelId": "entry-summary-v1"},
#     "PARTS WORKSHEET": {"status": "pending"}
#   }
# }
# ════════════════════════════════════════════════════════════════
def _set_stage(
    container,
    job_id: str,
    stage:  str,
    extra:  Optional[Dict] = None,
) -> None:
    """
    Patch the job document: update current stage, append to stagelog.
    Read-modify-write preserves all existing fields (extraction map, etc.).
    """
    try:
        now = utc_now_iso()
        try:
            doc = container.read_item(item=job_id, partition_key=job_id)
        except CosmosResourceNotFoundError:
            doc = {"id": job_id, "type": "bundle_result"}

        stagelog = doc.get("stagelog") or []
        stagelog.append({"stage": stage, "at": now})

        doc["stage"]    = stage
        doc["stageAt"]  = now
        doc["stagelog"] = stagelog

        if stage in (STAGE_COMPLETED, STAGE_FAILED, STAGE_DISCARDED):
            doc["status"] = stage
        elif stage != STAGE_QUEUED:
            doc["status"] = "processing"

        if extra:
            doc.update(extra)

        container.upsert_item(doc)
        logger.info("[STAGE] job=%s  →  %s", job_id, stage)
    except Exception as exc:
        logger.warning("[STAGE] Could not set stage=%s for job=%s: %s", stage, job_id, exc)


def _set_extraction_substatus(
    container,
    job_id:   str,
    doc_type: str,
    status:   str,
    extra:    Optional[Dict] = None,
) -> None:
    """Patch the per-doc-type extraction sub-status inside the job document."""
    try:
        now = utc_now_iso()
        try:
            doc = container.read_item(item=job_id, partition_key=job_id)
        except CosmosResourceNotFoundError:
            doc = {"id": job_id, "type": "bundle_result"}

        extraction = doc.get("extraction") or {}
        entry      = extraction.get(doc_type) or {}
        entry["status"] = status

        if status == SUB_PROCESSING:
            entry["startedAt"] = now
        elif status in (SUB_DONE, SUB_FAILED):
            entry["completedAt"] = now

        if extra:
            entry.update(extra)

        extraction[doc_type] = entry
        doc["extraction"]    = extraction
        container.upsert_item(doc)
        logger.info("[EXTRACT] job=%s  %s  →  %s", job_id, doc_type, status)
    except Exception as exc:
        logger.warning(
            "[EXTRACT] Could not set substatus job=%s docType=%s: %s",
            job_id, doc_type, exc,
        )


def _init_extraction_map(container, job_id: str, doc_types: List[str]) -> None:
    """Seed all detected doc types as 'pending' so UI shows the full picture immediately."""
    try:
        doc = container.read_item(item=job_id, partition_key=job_id)
        doc["extraction"] = {dt: {"status": SUB_PENDING} for dt in doc_types}
        container.upsert_item(doc)
    except Exception as exc:
        logger.warning("[EXTRACT] Could not init extraction map job=%s: %s", job_id, exc)


# ════════════════════════════════════════════════════════════════
# DI VALUE NORMALIZATION
# ════════════════════════════════════════════════════════════════
def _jsonable(obj: Any) -> Any:
    """Recursively convert any DI SDK object to a JSON-serialisable structure."""
    if obj is None:
        return None
    if isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, bytes):
        return base64.b64encode(obj).decode()
    if isinstance(obj, list):
        return [_jsonable(x) for x in obj]
    if isinstance(obj, dict):
        return {k: _jsonable(v) for k, v in obj.items()}
    for attr in ("to_dict", "as_dict", "model_dump"):
        fn = getattr(obj, attr, None)
        if callable(fn):
            try:
                return _jsonable(fn())
            except Exception:
                pass
    d = getattr(obj, "__dict__", None)
    if isinstance(d, dict) and d:
        return {k: _jsonable(v) for k, v in d.items()}
    return str(obj)


def _normalize_di_value(v: Any) -> Any:
    """Unwrap DI typed-value wrappers to plain Python scalars/structures."""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, list):
        return [_normalize_di_value(x) for x in v]
    if isinstance(v, dict):
        for key in (
            "valueString", "valueNumber", "valueInteger", "valueBoolean",
            "valueDate", "valueTime", "valuePhoneNumber", "valueSelectionMark",
        ):
            if key in v:
                return v[key]
        if "valueObject" in v and isinstance(v["valueObject"], dict):
            return {k: _normalize_di_value(val) for k, val in v["valueObject"].items()}
        if "valueArray" in v and isinstance(v["valueArray"], list):
            return [_normalize_di_value(i) for i in v["valueArray"]]
        drop = {"boundingRegions", "polygon", "spans", "content"}
        return {k: _normalize_di_value(val) for k, val in v.items() if k not in drop}
    return _normalize_di_value(_jsonable(v))


def _field_value(field: Any) -> Any:
    """Extract a plain value from a DI AnalyzedField object."""
    if field is None:
        return None
    value = getattr(field, "value", None)
    if value is not None:
        return _normalize_di_value(_jsonable(value))
    content = getattr(field, "content", None)
    if content is not None:
        return content
    return _normalize_di_value(_jsonable(field))


def _simplify_analyze_result(result: Any) -> Dict:
    """Return {fields: {...}, hasData: bool} from a DI AnalyzeResult."""
    docs = getattr(result, "documents", None) or []
    if not docs:
        return {"fields": {}, "hasData": False}
    raw    = getattr(docs[0], "fields", None) or {}
    simple = {str(k): _field_value(v) for k, v in raw.items()}
    return {
        "fields":  simple,
        "hasData": any(v not in (None, "", [], {}) for v in simple.values()),
    }


# ════════════════════════════════════════════════════════════════
# PARTS WORKSHEET — LINE ITEM NORMALIZATION
# ════════════════════════════════════════════════════════════════
def _unwrap_line_item_field(fv: Any) -> Any:
    """
    Fully unwrap a single DI field value to a plain scalar or structure.
    Handles all known DI response shapes including content-only fallback
    (low-confidence / rotated pages where valueString is absent).
    """
    if fv is None:
        return None
    if isinstance(fv, (str, int, float, bool)):
        return fv
    if isinstance(fv, list):
        return [_unwrap_line_item_field(i) for i in fv]
    if isinstance(fv, dict):
        for key in (
            "valueString", "valueNumber", "valueInteger", "valueDate",
            "valueTime", "valueBoolean", "valuePhoneNumber", "valueSelectionMark",
        ):
            if key in fv:
                val = fv[key]
                return _unwrap_line_item_field(val) if isinstance(val, dict) else val
        if "valueObject" in fv and isinstance(fv["valueObject"], dict):
            return {k: _unwrap_line_item_field(v) for k, v in fv["valueObject"].items()}
        if "valueArray" in fv and isinstance(fv["valueArray"], list):
            return [_unwrap_line_item_field(i) for i in fv["valueArray"]]
        # content-only fallback — DI omits valueString on low-confidence / rotated pages
        if "content" in fv:
            return str(fv["content"])
        # No usable value — return None rather than leaking a metadata dict
        return None
    return str(fv)


def _dedup_spacejoined(s: str) -> str:
    """
    DI sometimes returns 'VALUE VALUE' when a field spans multiple OCR regions.
    Deduplicates exact half-repetitions only — never mangles legitimate strings.
      'COOLER CORE COOLER CORE' → 'COOLER CORE'
      '4965482 4965483'         → '4965482, 4965483'  (distinct — left unchanged)
    """
    s = s.strip()
    if not s:
        return s
    tokens = s.split()
    n      = len(tokens)
    unique = list(dict.fromkeys(tokens))
    if len(unique) == 1:
        return unique[0]
    if n >= 2 and n % 2 == 0:
        h = n // 2
        if " ".join(tokens[:h]) == " ".join(tokens[h:]):
            return " ".join(tokens[:h])
    return s


def _fix_parts_worksheet_line_items(pw_fields: Dict) -> Dict:
    """
    Normalise parts_worksheet.LineItems to a plain list[dict].
    Handles three storage cases:
      A — already a clean list   (new docs after fix)
      B — DI array wrapper       (old docs before fix)
      C — missing / None
    Also applies _dedup_spacejoined to all string values.
    Safe to call on both freshly extracted and old Cosmos documents.
    """
    if not isinstance(pw_fields, dict):
        return pw_fields

    line_items = pw_fields.get("LineItems")
    raw: List[Any] = []

    if isinstance(line_items, list):
        raw = line_items
    elif isinstance(line_items, dict) and "valueArray" in line_items:
        for item in (line_items.get("valueArray") or []):
            if not isinstance(item, dict):
                continue
            vo = item.get("valueObject", {})
            raw.append(vo if isinstance(vo, dict) else item)
    else:
        pw_fields["LineItems"] = []
        return pw_fields

    clean: List[Dict] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        row: Dict = {}
        for k, v in item.items():
            unwrapped = _unwrap_line_item_field(v)
            if isinstance(unwrapped, str):
                unwrapped = _dedup_spacejoined(unwrapped)
            row[k] = unwrapped
        clean.append(row)

    pw_fields["LineItems"] = clean
    return pw_fields


# ════════════════════════════════════════════════════════════════
# PDF HELPERS
# ════════════════════════════════════════════════════════════════
def _extract_pages(reader: PdfReader, page_numbers: List[int]) -> bytes:
    """
    Render specific 1-based page numbers from an already-parsed PdfReader.
    Caller should parse PDF bytes once and reuse the reader object.
    """
    writer = PdfWriter()
    for p in page_numbers:
        writer.add_page(reader.pages[p - 1])
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════
# INVOICE DATE PARSING
# ════════════════════════════════════════════════════════════════
def _parse_date(value: Any) -> Optional[str]:
    """Try to parse any date-like value to YYYY-MM-DD. Returns None if unparseable."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Already ISO
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    s2    = s.replace(".", "/").replace("-", "/")
    parts = [p.strip() for p in s2.split("/") if p.strip()]
    # YYYY/MM/DD
    if len(parts) == 3 and len(parts[0]) == 4 and parts[0].isdigit():
        y, m, d = parts[0], parts[1].zfill(2), parts[2].zfill(2)
        if m.isdigit() and d.isdigit():
            return f"{y}-{m}-{d}"
    # DD/MM/YYYY or MM/DD/YYYY
    if len(parts) == 3 and len(parts[2]) == 4 and parts[2].isdigit():
        a, b, y = parts[0], parts[1], parts[2]
        if a.isdigit() and b.isdigit():
            da, db = int(a), int(b)
            d, m   = (da, db) if da > 12 else (db, da) if db > 12 else (da, db)
            return f"{y}-{str(m).zfill(2)}-{str(d).zfill(2)}"
    for fmt in ("%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(
                s.replace("-", " ").replace("/", " ").strip(), fmt
            ).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def _extract_ceva_invoice_date(doc: Dict) -> Optional[str]:
    ceva = doc.get("ceva") if isinstance(doc.get("ceva"), dict) else {}
    if CEVA_INVOICE_DATE_FIELD in ceva:
        parsed = _parse_date(ceva[CEVA_INVOICE_DATE_FIELD])
        if parsed:
            return parsed
    for k, v in ceva.items():
        lk = str(k).strip().lower()
        if "invoice" in lk and "date" in lk:
            parsed = _parse_date(v)
            if parsed:
                return parsed
    return None


def _compute_date_range(range_key: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not range_key:
        return None, None
    rk    = range_key.strip().lower()
    today = utc_now().date()
    if rk == "all":
        return None, None
    if rk in ("last_week", "last_7_days"):
        return (today - timedelta(days=7)).strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    if rk == "last_30_days":
        return (today - timedelta(days=30)).strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    if rk == "last_month":
        first_this = today.replace(day=1)
        last_prev  = first_this - timedelta(days=1)
        first_prev = last_prev.replace(day=1)
        return first_prev.strftime("%Y-%m-%d"), last_prev.strftime("%Y-%m-%d")
    return None, None


# ════════════════════════════════════════════════════════════════
# SUMMARY COLUMNS (shared by list + excel)
# ════════════════════════════════════════════════════════════════
SUMMARY_COLS: List[Tuple[str, str]] = [
    ("fileName",                   "File Name"),
    ("cevaInvoiceDate",            "CEVA INVOICE DATE"),
    ("invoiceNumber",              "INVOICE NUMBER"),
    ("entryNumber",                "ENTRY NUMBER"),
    ("departureDate",              "DEPARTURE DATE"),
    ("arrivalDate",                "ARRIVAL DATE"),
    ("descriptionOfGoods",         "DESCRIPTION OF GOODS"),
    ("totalEnteredValue",          "TOTAL ENTERED VALUE"),
    ("totalOtherFees",             "TOTAL OTHER FEES"),
    ("duty",                       "DUTY"),
    ("supplier",                   "SUPPLIER"),
    ("parts_Invoice_Seq",          "PARTS Invoice_Seq"),
    ("parts_Invoice_Number",       "PARTS Invoice_Number"),
    ("parts_Part_No",              "PARTS Part_No"),
    ("parts_HTS_code",             "PARTS HTS_code"),
    ("parts_Description_of_Goods", "PARTS Description_of_Goods"),
]

def _safe_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    try:
        return json.dumps(v, ensure_ascii=False)
    except Exception:
        return str(v)


def _join_line_item_values(line_items: List[Dict], field: str) -> str:
    """Collect field from all line items and join distinct non-empty values with ', '."""
    seen: List[str] = []
    for item in (line_items or []):
        v = item.get(field)
        if v is None:
            continue
        s = str(v).strip()
        if s and s not in seen:
            seen.append(s)
    return ", ".join(seen)


# ════════════════════════════════════════════════════════════════
# COSMOS QUERY BUILDER
# ════════════════════════════════════════════════════════════════
def _build_filters(req: func.HttpRequest) -> Tuple[str, List[Dict]]:
    q            = req.params.get("q")
    range_key    = req.params.get("range")
    invoice_from = _parse_iso_date_only(req.params.get("invoiceFrom"))
    invoice_to   = _parse_iso_date_only(req.params.get("invoiceTo"))

    if range_key and not (invoice_from or invoice_to) and range_key.strip().lower() != "all":
        invoice_from, invoice_to = _compute_date_range(range_key)

    where:  List[str]  = ['c.type = "bundle_result"', 'c.status = "completed"']
    params: List[Dict] = []

    if q:
        where.append("CONTAINS(LOWER(c.fileName), LOWER(@q))")
        params.append({"name": "@q", "value": q})
    if invoice_from:
        where.append("IS_DEFINED(c.cevaInvoiceDate) AND c.cevaInvoiceDate >= @invFrom")
        params.append({"name": "@invFrom", "value": invoice_from})
    if invoice_to:
        where.append("IS_DEFINED(c.cevaInvoiceDate) AND c.cevaInvoiceDate <= @invTo")
        params.append({"name": "@invTo", "value": invoice_to})

    return " AND ".join(where), params


# ════════════════════════════════════════════════════════════════
# HTTP: UPLOAD  POST /upload
# ════════════════════════════════════════════════════════════════
@app.route(route="upload", methods=["POST"])
def upload(req: func.HttpRequest) -> func.HttpResponse:
    files = req.files.getlist("files")
    if not files:
        return _json_response(
            {"error": "Upload files via multipart/form-data field 'files'"},
            status_code=400,
        )

    uploads = _blob_service().get_container_client(UPLOADS_CONTAINER)
    try:
        uploads.create_container()
    except Exception:
        pass

    # ── Order matters ─────────────────────────────────────────
    # 1. Register new batchId in Cosmos FIRST.
    #    Any already-running workers will read this and discard themselves.
    # 2. Clear visible queue messages (best effort).
    # 3. Enqueue new jobs — messages arrive AFTER the clear.
    batch_id = str(uuid.uuid4())
    try:
        _set_active_batch(batch_id)
    except Exception as exc:
        logger.error("[UPLOAD] Cannot set active batch — aborting: %s", exc)
        return _json_response(
            {"error": f"Storage error — could not register batch: {exc}"},
            status_code=500,
        )

    _clear_stale_queue_messages()  # best-effort, never raises

    qc   = _queue_client()
    jobs: List[Dict] = []

    def _enqueue(pdf_bytes: bytes, filename: str) -> Dict:
        job_id    = str(uuid.uuid4())
        blob_path = f"{job_id}/original/{filename}"
        now       = utc_now_iso()

        uploads.upload_blob(blob_path, pdf_bytes, overwrite=True)

        # Write initial Cosmos record immediately so status endpoint works right away
        _cosmos_container().upsert_item({
            "id":        job_id,
            "type":      "bundle_result",
            "fileName":  filename,
            "batchId":   batch_id,
            "createdAt": now,
            "status":    STAGE_QUEUED,
            "stage":     STAGE_QUEUED,
            "stageAt":   now,
            "stagelog":  [{"stage": STAGE_QUEUED, "at": now}],
        })

        msg = qc.send_message(json.dumps({
            "jobId":       job_id,
            "pdfBlobPath": blob_path,
            "fileName":    filename,
            "batchId":     batch_id,
        }))
        logger.info(
            "[ENQUEUE] jobId=%s  file=%s  msgId=%s  batch=%s",
            job_id, filename, msg.id, batch_id,
        )
        return {
            "jobId":    job_id,
            "fileName": filename,
            "blobPath": blob_path,
            "bytes":    len(pdf_bytes),
        }

    try:
        for f in files:
            filename = getattr(f, "filename", "file.bin")
            f.stream.seek(0)
            data = f.stream.read()

            if _is_zip(filename):
                with zipfile.ZipFile(io.BytesIO(data), "r") as zf:
                    for zi in zf.infolist():
                        if zi.is_dir() or not _is_pdf(zi.filename):
                            continue
                        leaf = zi.filename.split("/")[-1].split("\\")[-1] or "file.pdf"
                        jobs.append(_enqueue(zf.read(zi), leaf))
            elif _is_pdf(filename):
                jobs.append(_enqueue(data, filename))
            else:
                return _json_response(
                    {"error": f"Unsupported file: {filename}. Upload PDF or ZIP of PDFs."},
                    status_code=400,
                )

        resp: Dict = {"batchId": batch_id, "jobs": jobs}
        if len(jobs) == 1:
            resp["jobId"] = jobs[0]["jobId"]
        logger.info("[UPLOAD] batchId=%s  jobCount=%d", batch_id, len(jobs))
        return _json_response(resp)

    except zipfile.BadZipFile:
        return _json_response({"error": "Invalid ZIP file."}, status_code=400)
    except Exception as exc:
        logger.exception("[UPLOAD] Unexpected error: %s", exc)
        return _json_response({"error": str(exc)}, status_code=500)


# ════════════════════════════════════════════════════════════════
# QUEUE WORKER
# ════════════════════════════════════════════════════════════════
@app.function_name(name="job_worker")
@app.queue_trigger(arg_name="msg", queue_name="jobs", connection="AzureWebJobsStorage")
def job_worker(msg: func.QueueMessage) -> None:
    payload       = json.loads(msg.get_body().decode())
    job_id        = payload["jobId"]
    pdf_blob_path = payload["pdfBlobPath"]
    file_name     = payload["fileName"]
    msg_batch_id  = payload.get("batchId")

    uploads   = _blob_service().get_container_client(UPLOADS_CONTAINER)
    container = _cosmos_container()

    logger.info(
        "[WORKER] START  job=%s  file=%s  batch=%s",
        job_id, file_name, msg_batch_id,
    )

    # ── Staleness check ───────────────────────────────────────
    # If a newer upload arrived the active batch will differ.
    # Discard this job, clean up its blob, and return without processing.
    active_batch = _get_active_batch_id()
    if msg_batch_id and active_batch and msg_batch_id != active_batch:
        logger.info(
            "[WORKER] DISCARD  job=%s  msgBatch=%s  activeBatch=%s",
            job_id, msg_batch_id, active_batch,
        )
        _set_stage(container, job_id, STAGE_DISCARDED, {
            "discardReason": f"Superseded by batch {active_batch}",
            "discardedAt":   utc_now_iso(),
        })
        try:
            uploads.delete_blob(pdf_blob_path)
        except Exception as exc:
            logger.warning("[WORKER] Could not delete stale blob: %s", exc)
        return

    created_at = utc_now_iso()

    try:
        # ── Stage 1: classifying ──────────────────────────────
        _set_stage(container, job_id, STAGE_CLASSIFYING, {
            "fileName":  file_name,
            "createdAt": created_at,
        })
        logger.info("[WORKER] CLASSIFYING  job=%s", job_id)

        pdf_bytes = uploads.get_blob_client(pdf_blob_path).download_blob().readall()

        # Parse PDF once — the PdfReader is shared read-only across extraction threads
        reader = PdfReader(io.BytesIO(pdf_bytes))

        clf_result = _di_call_with_retry(
            fn=lambda: _di_client().begin_classify_document(
                CLASSIFIER_ID,
                body=pdf_bytes,
                content_type="application/pdf",
                split="auto",
            ),
            timeout_sec=DI_CLASSIFY_TIMEOUT_SEC,
            label=f"classify job={job_id}",
        )
        logger.info("[WORKER] CLASSIFIED  job=%s", job_id)

        best_docs: Dict[str, Dict] = {}
        for doc in clf_result.documents:
            dt    = _normalize_doctype(doc.doc_type)
            conf  = doc.confidence
            pages = [r.page_number for r in doc.bounding_regions]
            if dt in TARGET_DOC_TYPES:
                if not best_docs.get(dt) or conf > best_docs[dt]["confidence"]:
                    best_docs[dt] = {"confidence": conf, "pages": pages}

        logger.info(
            "[WORKER] Detected: %s  job=%s",
            list(best_docs.keys()), job_id,
        )

        # ── Stage 2: extracting ───────────────────────────────
        _set_stage(container, job_id, STAGE_EXTRACTING, {
            "detectedDocTypes": list(best_docs.keys()),
        })
        _init_extraction_map(container, job_id, list(best_docs.keys()))

        def _extract_one(doc_type: str, info: Dict) -> Optional[Dict]:
            """
            Run one DI extraction model for a specific doc type.
            Executed concurrently — each thread owns its own DI client.
            Updates Cosmos sub-status at start, completion, and failure.
            """
            model_id = EXTRACTION_MODELS.get(doc_type)
            if not model_id:
                logger.warning("[WORKER] No model for %s  job=%s", doc_type, job_id)
                return None

            _set_extraction_substatus(container, job_id, doc_type, SUB_PROCESSING, {
                "modelId": model_id,
                "pages":   info["pages"],
            })
            logger.info(
                "[WORKER] EXTRACTING  job=%s  docType=%s  pages=%s",
                job_id, doc_type, info["pages"],
            )

            try:
                extracted_pdf = _extract_pages(reader, info["pages"])
                result = _di_call_with_retry(
                    fn=lambda: _di_client().begin_analyze_document(
                        model_id,
                        body=extracted_pdf,
                        content_type="application/pdf",
                    ),
                    timeout_sec=DI_EXTRACT_TIMEOUT_SEC,
                    label=f"extract {doc_type} job={job_id}",
                )
                simplified = _simplify_analyze_result(result)
                fields     = simplified.get("fields", {})

                key = _safe_slug(doc_type).lower()
                if key == "parts_worksheet":
                    fields = _fix_parts_worksheet_line_items(fields)

                _set_extraction_substatus(container, job_id, doc_type, SUB_DONE, {
                    "fieldCount": len(fields),
                })
                logger.info(
                    "[WORKER] EXTRACTED  job=%s  docType=%s  fields=%d",
                    job_id, doc_type, len(fields),
                )
                return {
                    "doc_type":   doc_type,
                    "key":        key,
                    "model_id":   model_id,
                    "pages":      info["pages"],
                    "confidence": info["confidence"],
                    "fields":     fields,
                }

            except Exception as exc:
                logger.error(
                    "[WORKER] EXTRACT FAILED  job=%s  docType=%s  error=%s",
                    job_id, doc_type, exc,
                )
                _set_extraction_substatus(container, job_id, doc_type, SUB_FAILED, {
                    "error": str(exc),
                })
                return None  # Non-fatal — other doc types continue

        merged:         Dict[str, Any] = {}
        picked_pages:   Dict[str, Any] = {}
        confidence:     Dict[str, Any] = {}
        extracted_docs: List[Dict]     = []

        with concurrent.futures.ThreadPoolExecutor(max_workers=max(len(best_docs), 1)) as pool:
            future_map = {
                pool.submit(_extract_one, dt, info): dt
                for dt, info in best_docs.items()
            }
            for future in concurrent.futures.as_completed(future_map):
                result = future.result()
                if result is None:
                    continue
                merged[result["key"]]            = result["fields"]
                picked_pages[result["doc_type"]] = result["pages"]
                confidence[result["doc_type"]]   = result["confidence"]
                extracted_docs.append({
                    "docType":    result["doc_type"],
                    "modelId":    result["model_id"],
                    "pages":      result["pages"],
                    "fieldCount": len(result["fields"]),
                })

        # ── Stage 3: saving ───────────────────────────────────
        _set_stage(container, job_id, STAGE_SAVING)
        logger.info("[WORKER] SAVING  job=%s", job_id)

        # Preserve stagelog + extraction map accumulated during processing
        try:
            existing         = container.read_item(item=job_id, partition_key=job_id)
            saved_stagelog   = existing.get("stagelog") or []
            saved_extraction = existing.get("extraction") or {}
        except Exception:
            saved_stagelog   = []
            saved_extraction = {}

        completed_at = utc_now_iso()
        saved_stagelog.append({"stage": STAGE_COMPLETED, "at": completed_at})

        final_doc: Dict[str, Any] = {
            "id":                job_id,
            "type":              "bundle_result",
            "fileName":          file_name,
            "batchId":           msg_batch_id,
            "createdAt":         created_at,
            "completedAt":       completed_at,
            "status":            STAGE_COMPLETED,
            "stage":             STAGE_COMPLETED,
            "stageAt":           completed_at,
            "stagelog":          saved_stagelog,
            "extraction":        saved_extraction,
            "sourcePdfBlobPath": pdf_blob_path,
            "pickedPages":       picked_pages,
            "confidence":        confidence,
            "extractedDocs":     extracted_docs,
            **merged,
        }
        final_doc["cevaInvoiceDate"] = _extract_ceva_invoice_date(final_doc)

        container.upsert_item(final_doc)
        logger.info("[WORKER] COMPLETED  job=%s  file=%s", job_id, file_name)

        # Delete blob only after successful Cosmos write
        _delete_blob_prefix(uploads, f"{job_id}/")

    except Exception as exc:
        logger.exception("[WORKER] FAILED  job=%s  error=%s", job_id, exc)
        _set_stage(container, job_id, STAGE_FAILED, {
            "status":            STAGE_FAILED,
            "error":             str(exc),
            "failedAt":          utc_now_iso(),
            "fileName":          file_name,
            "createdAt":         created_at,
            "sourcePdfBlobPath": pdf_blob_path,
        })
        raise  # Re-raise so Azure Functions retries / dead-letters the message


# ════════════════════════════════════════════════════════════════
# HTTP: JOB STATUS  GET /job/{jobId}/status
#
# Lightweight polling endpoint — returns live progress detail.
# Poll every 2–3 seconds while stage not in {completed, failed, discarded}.
#
# Response shape:
# {
#   "id":       "abc-123",
#   "fileName": "shipment.pdf",
#   "stage":    "extracting",
#   "stageAt":  "...",
#   "status":   "processing",
#   "stagelog": [{"stage":"queued","at":"..."},{"stage":"classifying","at":"..."},...],
#   "detectedDocTypes": ["CEVA","ENTRY SUMMARY","PARTS WORKSHEET"],
#   "extraction": {
#     "CEVA":            {"status":"done",       "startedAt":"...","completedAt":"...","fieldCount":12},
#     "ENTRY SUMMARY":   {"status":"processing", "startedAt":"...","modelId":"entry-summary-v1"},
#     "PARTS WORKSHEET": {"status":"pending"}
#   }
# }
# ════════════════════════════════════════════════════════════════
@app.route(route="job/{jobId}/status", methods=["GET"])
def get_job_status(req: func.HttpRequest) -> func.HttpResponse:
    job_id = req.route_params.get("jobId")
    if not job_id:
        return _json_response({"error": "Missing jobId"}, status_code=400)

    try:
        doc = _cosmos_container().read_item(item=job_id, partition_key=job_id)
    except CosmosResourceNotFoundError:
        return _json_response(
            {"id": job_id, "stage": "not_found", "status": "not_found"},
            status_code=404,
        )

    out: Dict[str, Any] = {
        "id":               doc.get("id"),
        "fileName":         doc.get("fileName"),
        "batchId":          doc.get("batchId"),
        "stage":            doc.get("stage") or doc.get("status"),
        "stageAt":          doc.get("stageAt"),
        "status":           doc.get("status"),
        "createdAt":        doc.get("createdAt"),
        "stagelog":         doc.get("stagelog") or [],
        "detectedDocTypes": doc.get("detectedDocTypes") or [],
        "extraction":       doc.get("extraction") or {},
    }
    for key in ("completedAt", "error", "failedAt", "discardReason", "discardedAt"):
        if doc.get(key):
            out[key] = doc[key]

    # ── Stuck detection ───────────────────────────────────────
    # If a job has been in classifying/extracting/saving for >5 min
    # without advancing, flag it so the UI can offer a retry.
    STUCK_THRESHOLD_SEC = 300
    stage = doc.get("stage") or ""
    stage_at_str = doc.get("stageAt")
    if stage in (STAGE_CLASSIFYING, STAGE_EXTRACTING, STAGE_SAVING) and stage_at_str:
        try:
            stage_at = datetime.fromisoformat(stage_at_str)
            if stage_at.tzinfo is None:
                stage_at = stage_at.replace(tzinfo=timezone.utc)
            elapsed = (utc_now() - stage_at).total_seconds()
            if elapsed > STUCK_THRESHOLD_SEC:
                out["stuck"]          = True
                out["stuckSec"]       = int(elapsed)
                out["unstickUrl"]     = f"/api/job/{job_id}/unstick"
        except Exception:
            pass

    return _json_response(out)


# ════════════════════════════════════════════════════════════════
# HTTP: UNSTICK  POST /job/{jobId}/unstick
#
# Re-queues a job that is stuck in classifying/extracting/saving.
# Reuses the original blob path so no re-upload is needed.
# ════════════════════════════════════════════════════════════════
@app.route(route="job/{jobId}/unstick", methods=["POST"])
def unstick_job(req: func.HttpRequest) -> func.HttpResponse:
    job_id = req.route_params.get("jobId")
    if not job_id:
        return _json_response({"error": "Missing jobId"}, status_code=400)

    try:
        container = _cosmos_container()
        doc       = container.read_item(item=job_id, partition_key=job_id)
    except CosmosResourceNotFoundError:
        return _json_response({"error": "Job not found"}, status_code=404)

    stage = doc.get("stage") or doc.get("status")
    if stage in (STAGE_COMPLETED, STAGE_DISCARDED):
        return _json_response(
            {"error": f"Cannot unstick a job in stage '{stage}'"},
            status_code=400,
        )

    blob_path = doc.get("sourcePdfBlobPath")
    file_name = doc.get("fileName")
    batch_id  = doc.get("batchId")

    if not blob_path:
        return _json_response(
            {"error": "No source blob path on record — cannot re-queue"},
            status_code=400,
        )

    # Verify the blob still exists before re-queuing
    try:
        _blob_service().get_container_client(UPLOADS_CONTAINER) \
            .get_blob_client(blob_path).get_blob_properties()
    except Exception:
        return _json_response(
            {"error": "Source blob no longer exists — original PDF was already cleaned up. Re-upload the file."},
            status_code=409,
        )

    now = utc_now_iso()

    # Reset stage to queued and re-queue
    stagelog = doc.get("stagelog") or []
    stagelog.append({"stage": "unstick_requested", "at": now})
    stagelog.append({"stage": STAGE_QUEUED,         "at": now})

    container.upsert_item({
        **doc,
        "status":   STAGE_QUEUED,
        "stage":    STAGE_QUEUED,
        "stageAt":  now,
        "stagelog": stagelog,
        "unstuckAt": now,
    })

    msg = _queue_client().send_message(json.dumps({
        "jobId":       job_id,
        "pdfBlobPath": blob_path,
        "fileName":    file_name,
        "batchId":     batch_id,
    }))
    logger.info("[UNSTICK] Re-queued job=%s  msgId=%s", job_id, msg.id)

    return _json_response({
        "jobId":     job_id,
        "fileName":  file_name,
        "stage":     STAGE_QUEUED,
        "messageId": msg.id,
        "unstuckAt": now,
    })


# ════════════════════════════════════════════════════════════════
# HTTP: BATCH STATUS  GET /batch/{batchId}/status
#
# See ALL jobs in a batch at once.
# Response includes a summary count per stage and allDone flag.
# ════════════════════════════════════════════════════════════════
@app.route(route="batch/{batchId}/status", methods=["GET"])
def get_batch_status(req: func.HttpRequest) -> func.HttpResponse:
    batch_id = req.route_params.get("batchId")
    if not batch_id:
        return _json_response({"error": "Missing batchId"}, status_code=400)

    query = """
    SELECT
      c.id, c.fileName, c.stage, c.stageAt, c.status,
      c.createdAt, c.completedAt, c.failedAt, c.discardedAt,
      c.stagelog, c.detectedDocTypes, c.extraction,
      c.error, c.discardReason
    FROM c
    WHERE c.batchId    = @batchId
      AND c.type       = "bundle_result"
      AND c.id        != @sentinelId
    """
    try:
        items = list(_cosmos_container().query_items(
            query=query,
            parameters=[
                {"name": "@batchId",    "value": batch_id},
                {"name": "@sentinelId", "value": ACTIVE_BATCH_DOC_ID},
            ],
            enable_cross_partition_query=True,
        ))

        summary: Dict[str, int] = {
            STAGE_QUEUED: 0, STAGE_CLASSIFYING: 0, STAGE_EXTRACTING: 0,
            STAGE_SAVING: 0, STAGE_COMPLETED: 0, STAGE_FAILED: 0, STAGE_DISCARDED: 0,
        }
        for item in items:
            stage = item.get("stage") or item.get("status") or "unknown"
            summary[stage] = summary.get(stage, 0) + 1

        terminal = {STAGE_COMPLETED, STAGE_FAILED, STAGE_DISCARDED}
        all_done = bool(items) and all(
            (item.get("stage") or item.get("status")) in terminal
            for item in items
        )

        return _json_response({
            "batchId": batch_id,
            "total":   len(items),
            "allDone": all_done,
            "summary": summary,
            "jobs":    items,
        })

    except Exception as exc:
        logger.exception("[BATCH STATUS] error: %s", exc)
        return _json_response({"error": str(exc)}, status_code=500)


# ════════════════════════════════════════════════════════════════
# HTTP: JOB DETAIL  GET /job/{jobId}/ui
# Returns full extracted fields once stage = completed.
# While still processing, returns current progress (same as /status).
# ════════════════════════════════════════════════════════════════
@app.route(route="job/{jobId}/ui", methods=["GET"])
def get_job_ui(req: func.HttpRequest) -> func.HttpResponse:
    job_id = req.route_params.get("jobId")
    if not job_id:
        return _json_response({"error": "Missing jobId"}, status_code=400)

    try:
        doc = _cosmos_container().read_item(item=job_id, partition_key=job_id)
    except CosmosResourceNotFoundError:
        return _json_response({"id": job_id, "status": "not_found"}, status_code=404)

    status = doc.get("status")

    def _strip(d: Any) -> Dict:
        return {k: v for k, v in d.items() if not k.startswith("_")} if isinstance(d, dict) else {}

    if status != STAGE_COMPLETED:
        # Return progress info while job is in-flight or has failed
        out: Dict[str, Any] = {
            "id":               doc.get("id"),
            "fileName":         doc.get("fileName"),
            "status":           status,
            "stage":            doc.get("stage") or status,
            "stageAt":          doc.get("stageAt"),
            "createdAt":        doc.get("createdAt"),
            "stagelog":         doc.get("stagelog") or [],
            "extraction":       doc.get("extraction") or {},
            "detectedDocTypes": doc.get("detectedDocTypes") or [],
        }
        for key in ("error", "failedAt", "discardReason", "discardedAt"):
            if doc.get(key):
                out[key] = doc[key]
        return _json_response(out)

    # Completed — return full extracted fields
    # Fix LineItems on old docs without requiring reprocessing
    pw_fixed = _fix_parts_worksheet_line_items(_strip(doc.get("parts_worksheet") or {}))

    return _json_response({
        "id":              doc.get("id"),
        "fileName":        doc.get("fileName"),
        "status":          STAGE_COMPLETED,
        "stage":           STAGE_COMPLETED,
        "createdAt":       doc.get("createdAt"),
        "completedAt":     doc.get("completedAt"),
        "cevaInvoiceDate": doc.get("cevaInvoiceDate"),
        "stagelog":        doc.get("stagelog") or [],
        "extraction":      doc.get("extraction") or {},
        "ceva":            _strip(doc.get("ceva") or {}),
        "entry_summary":   _strip(doc.get("entry_summary") or {}),
        "parts_worksheet": pw_fixed,
    })


# ════════════════════════════════════════════════════════════════
# HTTP: RESULTS LIST  GET /results
# Paginated list of completed jobs with summary columns.
# ════════════════════════════════════════════════════════════════
@app.route(route="results", methods=["GET"])
def list_results(req: func.HttpRequest) -> func.HttpResponse:
    container = _cosmos_container()
    page_size = max(1, min(MAX_PAGE_SIZE, _parse_int(req.params.get("pageSize"), DEFAULT_PAGE_SIZE)))
    token     = req.params.get("token")
    where_clause, params = _build_filters(req)

    query = f"""
    SELECT
      c.id,
      c.fileName,
      c.cevaInvoiceDate,
      c.ceva["INVOICE NUMBER"]       AS invoiceNumber,
      c.ceva["ENTRY NUMBER"]         AS entryNumber,
      c.ceva["DEPARTURE DATE"]       AS departureDate,
      c.ceva["ARRIVAL DATE"]         AS arrivalDate,
      c.ceva["DESCRIPTION OF GOODS"] AS descriptionOfGoods,
      c.entry_summary.Total_Entered_Value AS totalEnteredValue,
      c.entry_summary.Total_Other_Fees    AS totalOtherFees,
      c.entry_summary.Duty                AS duty,
      c.parts_worksheet.Supplier          AS supplier,
      c.parts_worksheet.LineItems         AS lineItems
    FROM c
    WHERE {where_clause}
    ORDER BY c.cevaInvoiceDate DESC
    """

    try:
        iterable   = container.query_items(
            query=query, parameters=params,
            enable_cross_partition_query=True, max_item_count=page_size,
        )
        page_iter  = iterable.by_page(continuation_token=token)
        page       = next(page_iter)
        raw_items  = list(page)
        next_token = getattr(page, "continuation_token", None)

        items: List[Dict] = []
        for item in raw_items:
            li_raw     = item.pop("lineItems", None) or []
            line_items = _fix_parts_worksheet_line_items({"LineItems": li_raw}).get("LineItems") or []
            item["parts_Invoice_Seq"]          = _join_line_item_values(line_items, "Invoice_Seq")
            item["parts_Invoice_Number"]       = _join_line_item_values(line_items, "Invoice_Number")
            item["parts_Part_No"]              = _join_line_item_values(line_items, "Part_No")
            item["parts_HTS_code"]             = _join_line_item_values(line_items, "HTS_code")
            item["parts_Description_of_Goods"] = _join_line_item_values(line_items, "Description_of_Goods")
            items.append(item)

        return _json_response({"items": items, "continuationToken": next_token})

    except StopIteration:
        return _json_response({"items": [], "continuationToken": None})
    except Exception as exc:
        logger.exception("[RESULTS] error: %s", exc)
        return _json_response({"error": str(exc)}, status_code=500)


# ════════════════════════════════════════════════════════════════
# HTTP: EXCEL EXPORT  GET /results/excel
# Two sheets: Summary (joined values) + LineItems (one row each).
# ════════════════════════════════════════════════════════════════
def _autosize_columns(ws, max_col: int, max_width: int = 60) -> None:
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len    = max(
            (len(str(c.value)) for c in ws[col_letter] if c.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


@app.route(route="results/excel", methods=["GET"])
def export_results_excel(req: func.HttpRequest) -> func.HttpResponse:
    where_clause, params = _build_filters(req)

    query = f"""
    SELECT c.id, c.fileName, c.cevaInvoiceDate, c.ceva, c.entry_summary, c.parts_worksheet
    FROM c
    WHERE {where_clause}
    ORDER BY c.cevaInvoiceDate DESC
    """

    try:
        all_docs: List[Dict] = [
            doc
            for page in _cosmos_container().query_items(
                query=query, parameters=params,
                enable_cross_partition_query=True, max_item_count=200,
            ).by_page()
            for doc in page
        ]

        wb       = Workbook()
        ws       = wb.active
        ws.title = "Summary"
        ws.append([label for _, label in SUMMARY_COLS])

        ws_li = wb.create_sheet("LineItems")
        ws_li.append([
            "File Name", "CEVA INVOICE DATE",
            "Invoice_Seq", "Invoice_Number", "Part_No", "HTS_code", "Description_of_Goods",
        ])

        for d in all_docs:
            ceva = d.get("ceva")          if isinstance(d.get("ceva"),          dict) else {}
            es   = d.get("entry_summary") if isinstance(d.get("entry_summary"), dict) else {}
            pw   = _fix_parts_worksheet_line_items(dict(d.get("parts_worksheet") or {}))
            lis  = pw.get("LineItems") or []

            row_data: Dict[str, Any] = {
                "fileName":                    d.get("fileName"),
                "cevaInvoiceDate":             d.get("cevaInvoiceDate"),
                "invoiceNumber":               ceva.get("INVOICE NUMBER"),
                "entryNumber":                 ceva.get("ENTRY NUMBER"),
                "departureDate":               ceva.get("DEPARTURE DATE"),
                "arrivalDate":                 ceva.get("ARRIVAL DATE"),
                "descriptionOfGoods":          ceva.get("DESCRIPTION OF GOODS"),
                "totalEnteredValue":           es.get("Total_Entered_Value"),
                "totalOtherFees":              es.get("Total_Other_Fees"),
                "duty":                        es.get("Duty"),
                "supplier":                    pw.get("Supplier"),
                "parts_Invoice_Seq":           _join_line_item_values(lis, "Invoice_Seq"),
                "parts_Invoice_Number":        _join_line_item_values(lis, "Invoice_Number"),
                "parts_Part_No":               _join_line_item_values(lis, "Part_No"),
                "parts_HTS_code":              _join_line_item_values(lis, "HTS_code"),
                "parts_Description_of_Goods":  _join_line_item_values(lis, "Description_of_Goods"),
            }
            ws.append([_safe_cell(row_data.get(key)) for key, _ in SUMMARY_COLS])

            for li in lis:
                if not isinstance(li, dict):
                    continue
                ws_li.append([
                    _safe_cell(d.get("fileName")),
                    _safe_cell(d.get("cevaInvoiceDate")),
                    _safe_cell(li.get("Invoice_Seq")),
                    _safe_cell(li.get("Invoice_Number")),
                    _safe_cell(li.get("Part_No")),
                    _safe_cell(li.get("HTS_code")),
                    _safe_cell(li.get("Description_of_Goods")),
                ])

        _autosize_columns(ws,    len(SUMMARY_COLS))
        _autosize_columns(ws_li, 7)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        range_key    = (req.params.get("range") or "all").strip()
        invoice_from = _parse_iso_date_only(req.params.get("invoiceFrom"))
        invoice_to   = _parse_iso_date_only(req.params.get("invoiceTo"))
        suffix = (
            f"{invoice_from or '...'}_to_{invoice_to or '...'}"
            if (invoice_from or invoice_to)
            else range_key.lower()
        )
        filename = f"bundle_export_{suffix}.xlsx"

        return func.HttpResponse(
            body=out.getvalue(),
            status_code=200,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except Exception as exc:
        logger.exception("[EXCEL] error: %s", exc)
        return _json_response({"error": str(exc)}, status_code=500)
